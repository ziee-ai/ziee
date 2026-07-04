#!/usr/bin/env python3
"""
Generate the binary + text assets referenced by showcase.sql.

Every file here has a FIXED, deterministic filename that matches a `files`
row in showcase.sql (keyed by the UUIDs in ids.env). Re-running this script
is idempotent — it just overwrites the bytes.

Outputs land in ./files/ next to this script. The load script (load.sh) then
copies each into the running server's file-store at:

    <app_data>/files/originals/<owner_user_id>/<file_id>.<ext>

Deps: Pillow (PNG/JPG), openpyxl (xlsx). PDF/CSV/text are written by hand so
no extra deps are needed for them.

To add a NEW file case:
  1. add a generator function below + call it in main(),
  2. add a matching `files` row + content block in showcase.sql,
  3. map its <file_id>.<ext> in load.sh's FILE_MAP.
"""
import os
import struct
import zlib

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "files")
os.makedirs(OUT, exist_ok=True)


def out(name):
    return os.path.join(OUT, name)


# ---------------------------------------------------------------------------
# chart.png — a small bar chart drawn with Pillow (stands in for a
# code_sandbox-generated matplotlib artifact returned via resource_link).
# ---------------------------------------------------------------------------
def gen_chart_png():
    from PIL import Image, ImageDraw

    W, H = 640, 400
    img = Image.new("RGB", (W, H), "#0f172a")
    d = ImageDraw.Draw(img)
    d.text((20, 12), "Sales by Quarter (showcase chart.png)", fill="#e2e8f0")
    bars = [("Q1", 120, "#38bdf8"), ("Q2", 200, "#34d399"),
            ("Q3", 160, "#fbbf24"), ("Q4", 280, "#f472b6")]
    base_y, x = 360, 80
    for label, val, color in bars:
        d.rectangle([x, base_y - val, x + 90, base_y], fill=color)
        d.text((x + 30, base_y + 8), label, fill="#cbd5e1")
        d.text((x + 20, base_y - val - 16), str(val), fill="#cbd5e1")
        x += 130
    img.save(out("chart.png"), "PNG")


# ---------------------------------------------------------------------------
# photo.jpg — a JPEG gradient (exercises the image renderer with a lossy type).
# ---------------------------------------------------------------------------
def gen_photo_jpg():
    from PIL import Image

    W, H = 480, 320
    img = Image.new("RGB", (W, H))
    px = img.load()
    for y in range(H):
        for x in range(W):
            px[x, y] = (int(255 * x / W), int(255 * y / H), 128)
    img.save(out("photo.jpg"), "JPEG", quality=85)


# ---------------------------------------------------------------------------
# workbook.xlsx — 3 sheets, to exercise the XlsxBody multi-sheet Tabs renderer.
# ---------------------------------------------------------------------------
def gen_workbook_xlsx():
    from openpyxl import Workbook

    wb = Workbook()
    s1 = wb.active
    s1.title = "Summary"
    s1.append(["Metric", "Value", "Delta"])
    s1.append(["Revenue", 128000, "+12%"])
    s1.append(["Costs", 74000, "-3%"])
    s1.append(["Margin", 54000, "+21%"])

    s2 = wb.create_sheet("Regions")
    s2.append(["Region", "Q1", "Q2", "Q3", "Q4"])
    for r in [("NA", 40, 55, 60, 90), ("EU", 30, 45, 40, 70),
              ("APAC", 50, 60, 60, 80)]:
        s2.append(list(r))

    s3 = wb.create_sheet("Raw")
    s3.append(["id", "ts", "event", "amount"])
    for i in range(1, 26):
        s3.append([i, f"2026-07-{i:02d}", "purchase", i * 3.5])

    wb.save(out("workbook.xlsx"))


# ---------------------------------------------------------------------------
# data.csv — plain CSV for the CSV renderer.
# ---------------------------------------------------------------------------
def gen_data_csv():
    rows = ["gene,chromosome,expression,p_value",
            "TP53,17,8.42,0.0001",
            "EGFR,7,6.10,0.0034",
            "BRCA1,17,4.75,0.0210",
            "MYC,8,9.88,0.0000",
            "PTEN,10,3.21,0.0450"]
    with open(out("data.csv"), "w") as f:
        f.write("\n".join(rows) + "\n")


# ---------------------------------------------------------------------------
# report.pdf — a minimal but valid single-page PDF (hand-written, no deps).
# ---------------------------------------------------------------------------
def gen_report_pdf():
    text = "Showcase Report (report.pdf) - renders in the PDF viewer."
    content = f"BT /F1 18 Tf 72 720 Td ({text}) Tj ET".encode("latin-1")
    objs = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
        b"/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(content)).encode() + b" >>\nstream\n"
        + content + b"\nendstream",
    ]
    pdf = b"%PDF-1.4\n"
    offsets = []
    for i, o in enumerate(objs, start=1):
        offsets.append(len(pdf))
        pdf += f"{i} 0 obj\n".encode() + o + b"\nendobj\n"
    xref_pos = len(pdf)
    pdf += b"xref\n0 " + str(len(objs) + 1).encode() + b"\n"
    pdf += b"0000000000 65535 f \n"
    for off in offsets:
        pdf += f"{off:010d} 00000 n \n".encode()
    pdf += (b"trailer\n<< /Size " + str(len(objs) + 1).encode()
            + b" /Root 1 0 R >>\nstartxref\n"
            + str(xref_pos).encode() + b"\n%%EOF")
    with open(out("report.pdf"), "wb") as f:
        f.write(pdf)


# ---------------------------------------------------------------------------
# script.py — a code file (exercises the code/text file viewer).
# ---------------------------------------------------------------------------
def gen_script_py():
    src = '''#!/usr/bin/env python3
"""Attached code file — exercises the source-code file viewer."""


def fib(n: int) -> int:
    a, b = 0, 1
    for _ in range(n):
        a, b = b, a + b
    return a


if __name__ == "__main__":
    print([fib(i) for i in range(10)])
'''
    with open(out("script.py"), "w") as f:
        f.write(src)


# ---------------------------------------------------------------------------
# notes.md — a markdown file attachment.
# ---------------------------------------------------------------------------
def gen_notes_md():
    md = """# Project Notes (notes.md attachment)

- **Goal:** exercise the markdown *file* viewer (distinct from inline chat md).
- Supports `inline code`, [links](https://example.com), and tables:

| Step | Owner | Status |
|------|-------|--------|
| Spec | A     | done   |
| Impl | B     | wip    |
"""
    with open(out("notes.md"), "w") as f:
        f.write(md)


# ---------------------------------------------------------------------------
# large.txt — a big-ish text blob to test scrolling / truncation.
# ---------------------------------------------------------------------------
def gen_large_txt():
    lines = [f"{i:05d}  Lorem ipsum dolor sit amet, consectetur adipiscing "
             f"elit, sed do eiusmod tempor incididunt ut labore." for i in range(1, 801)]
    with open(out("large.txt"), "w") as f:
        f.write("\n".join(lines) + "\n")


def main():
    gen_chart_png()
    gen_photo_jpg()
    gen_workbook_xlsx()
    gen_data_csv()
    gen_report_pdf()
    gen_script_py()
    gen_notes_md()
    gen_large_txt()
    print("Generated files in", OUT)
    for n in sorted(os.listdir(OUT)):
        print(f"  {n:16} {os.path.getsize(os.path.join(OUT, n)):>8} bytes")


if __name__ == "__main__":
    main()
